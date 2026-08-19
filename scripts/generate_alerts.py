import argparse
from datetime import date, timedelta
from pathlib import Path
import random
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "generated" / "alerts-ds-generated.xlsx"

SYSTEM_AREA_GROUPS = {
    "Machine Control": "Control & Interface",
    "Connectivity": "Control & Interface",
    "User Interface": "Control & Interface",
    "Heating & Boiler": "Beverage Systems",
    "Water": "Beverage Systems",
    "Dispensing": "Beverage Systems",
    "Ingredients": "Supply & Payment",
    "Payment": "Supply & Payment",
    "Cleaning & Waste": "Operations & Safety",
    "Safety & Access": "Operations & Safety",
    "Cooling": "Operations & Safety",
    "Power": "Operations & Safety",
}

SYSTEM_AREA_ALIASES = {
    "Control Board": "Machine Control",
    "Main Controller": "Machine Control",
    "Connectivity Module": "Connectivity",
    "Network Interface": "Connectivity",
    "Touchscreen Display": "User Interface",
    "Receipt Printer": "User Interface",
    "Boiler and Tank Assembly": "Heating & Boiler",
    "Boiler Control": "Heating & Boiler",
    "Boiler Chamber": "Heating & Boiler",
    "Boiler and Heating Circuit": "Heating & Boiler",
    "Water Supply Line": "Water",
    "Water Tank": "Water",
    "Water Reservoir": "Water",
    "Dispense Nozzle": "Dispensing",
    "Pump Motor": "Dispensing",
    "Dispense Valve": "Dispensing",
    "Cup Dispenser": "Dispensing",
    "Pump Assembly": "Dispensing",
    "Brew Group": "Dispensing",
    "Coffee Grinder": "Dispensing",
    "Dispense Pump": "Dispensing",
    "Ingredient Hoppers": "Ingredients",
    "Bean Hopper": "Ingredients",
    "Milk Reservoir": "Ingredients",
    "Milk Pump": "Ingredients",
    "Payment Terminal": "Payment",
    "Payment Gateway Interface": "Payment",
    "Card Reader": "Payment",
    "Payment Reader Board": "Payment",
    "Cleaning System": "Cleaning & Waste",
    "Drip Tray": "Cleaning & Waste",
    "Service Scheduler": "Cleaning & Waste",
    "Door Interlock": "Safety & Access",
    "Cooling Fan": "Cooling",
    "Power Supply Unit": "Power",
}


def system_area_taxonomy(component):
    """Return the canonical system area and its major organizational group."""
    system_area = SYSTEM_AREA_ALIASES[component]
    return system_area, SYSTEM_AREA_GROUPS[system_area]


info_templates = [
    {
        "title": "Machine Startup Complete",
        "description": "Machine completed a normal startup sequence without faults.",
        "component": "Control Board",
        "service": "No service action required; the machine is operating within normal parameters.",
        "technician": "Log the event and confirm normal startup values remain stable across the next operating cycle.",
    },
    {
        "title": "Backend Heartbeat Confirmed",
        "description": "The device is successfully reporting to the service backend.",
        "component": "Connectivity Module",
        "service": "Remote monitoring connection is healthy and no intervention is needed.",
        "technician": "Review heartbeat timestamps and confirm they are within the expected interval for this site.",
    },
    {
        "title": "Scheduled Cleaning Cycle Logged",
        "description": "The maintenance cleaning cycle completed without interruption.",
        "component": "Cleaning System",
        "service": "Cleaning cycle completed successfully and the unit remained in service.",
        "technician": "Verify the cleaning log and confirm the next maintenance cycle is scheduled correctly.",
    },
    {
        "title": "Ingredient Inventory Balanced OK",
        "description": "Ingredient and product inventory status is within expected limits.",
        "component": "Ingredient Hoppers",
        "service": "Inventory counts are aligned with the machine configuration and no refill is needed.",
        "technician": "Cross-check inventory counters against the physical stock levels for reporting accuracy.",
    },
    {
        "title": "Firmware Version Check Passed",
        "description": "Current firmware version matches the approved release.",
        "component": "Main Controller",
        "service": "No patch or rollback is required at this time.",
        "technician": "Archive the firmware record and validate version alignment against the approved release list.",
    },
    {
        "title": "Payment Self Test Completed OK",
        "description": "The cashless payment reader completed its self-diagnostic routine successfully.",
        "component": "Payment Terminal",
        "service": "Cashless device is fit for customer transactions and no service response is needed.",
        "technician": "Retain the diagnostic record and monitor for any future reader instability.",
    },
    {
        "title": "Tank Temperature Stable",
        "description": "Beverage tank temperature remains within the defined operating band.",
        "component": "Boiler and Tank Assembly",
        "service": "Temperature regulation is stable and no user-facing issue is present.",
        "technician": "Check the last thermal trend and confirm no drift is emerging from previous service activity.",
    },
    {
        "title": "Water Quality Reading Normal",
        "description": "Water quality readings are within the expected operational range.",
        "component": "Water Supply Line",
        "service": "Water conditions are acceptable and no service escalation is necessary.",
        "technician": "Note the water quality trend for reference during the next preventive inspection.",
    },
    {
        "title": "Nozzle Purge Cycle Successful",
        "description": "The dispense nozzle purge cycle completed without blockage or abnormal flow.",
        "component": "Dispense Nozzle",
        "service": "The purge cycle completed successfully and the machine remains ready for service.",
        "technician": "Log the nozzle purge outcome and confirm flow metrics remain within expected values.",
    },
    {
        "title": "Door Sensor State Normal",
        "description": "Door interlock and access sensors are in the normal state.",
        "component": "Door Interlock",
        "service": "No security or access event is active and the machine is operating as expected.",
        "technician": "Monitor sensor state during the next operating cycle to confirm the latch remains stable.",
    },
    {
        "title": "Payment Transaction Completed OK",
        "description": "A recent customer transaction ended normally with no duplicate settlement event.",
        "component": "Payment Gateway Interface",
        "service": "The transaction lifecycle completed correctly with no response anomaly.",
        "technician": "Confirm settlement logs match the backend report and archive the transaction record.",
    },
    {
        "title": "Receipt Printer Queue Cleared",
        "description": "Receipt printing queue is empty and journal activity is normal.",
        "component": "Receipt Printer",
        "service": "Printer queue was cleared successfully and no paper or jam issue is present.",
        "technician": "Inspect remaining paper stock and confirm print quality remains acceptable.",
    },
]

warning_templates = [
    {
        "title": "Bean Hopper Level Low",
        "description": "The coffee bean hopper level is approaching the minimum threshold for continuous service.",
        "component": "Bean Hopper",
        "severity": "Sev4",
        "service": "Refill is recommended before the next customer peak period to avoid service disruption.",
        "technician": "Top up the hopper and confirm the level sensor reads correctly after refill.",
    },
    {
        "title": "Milk Tank Below Refill Target",
        "description": "Milk inventory is beneath the preferred refill limit but the machine remains operational.",
        "component": "Milk Reservoir",
        "severity": "Sev4",
        "service": "The unit is still serving beverages, but milk supply should be replenished soon.",
        "technician": "Refill the container and recalibrate the float sensor if the reading remains offset.",
    },
    {
        "title": "Water Tank Level Warning",
        "description": "The water tank is nearing the low-level threshold and may require replenishment soon.",
        "component": "Water Tank",
        "severity": "Sev4",
        "service": "The machine can continue operating, but refill should be scheduled before the next service window.",
        "technician": "Fill the tank and inspect the level sensor for a delayed or sticky float response.",
    },
    {
        "title": "Payment Retry Rate Elevated",
        "description": "The payment system is showing a slightly higher-than-normal retry rate.",
        "component": "Card Reader",
        "severity": "Sev4",
        "service": "Transactions are still completing, but card acceptance should be monitored for intermittent faults.",
        "technician": "Clean the reader slot and verify contactless and contact reads are consistently accepted.",
    },
    {
        "title": "Touchscreen Response Running Slow",
        "description": "The HMI touch panel is responding slower than the expected baseline.",
        "component": "Touchscreen Display",
        "severity": "Sev4",
        "service": "Customer interaction remains possible, but responsiveness is degrading and should be corrected.",
        "technician": "Clean the panel and recalibrate the touch matrix to restore expected response speed.",
    },
    {
        "title": "Drip Tray Nearly Full",
        "description": "The drip tray sensor reports the tray level approaching the service threshold.",
        "component": "Drip Tray",
        "severity": "Sev4",
        "service": "The machine remains available, but the tray should be emptied before the next operating cycle.",
        "technician": "Empty the tray, reset the sensor count, and verify the tray is reporting empty correctly.",
    },
    {
        "title": "Pump Start Delay Notice",
        "description": "A soft start delay was observed before beverage dispensing began.",
        "component": "Pump Motor",
        "severity": "Sev4",
        "service": "Dispense still works, but delivery timing is slower than expected and should be reviewed.",
        "technician": "Inspect the pump startup sequence and validate the control board is not limiting power unnecessarily.",
    },
    {
        "title": "Recipe Dispense Volume Low",
        "description": "One beverage recipe is dispensing slightly below the configured target volume.",
        "component": "Dispense Valve",
        "severity": "Sev4",
        "service": "Service is still available, but the recipe volume should be recalibrated during the next visit.",
        "technician": "Recalibrate the recipe volume and test three consecutive servings before returning the unit.",
    },
    {
        "title": "Machine Heartbeat Report Delayed",
        "description": "The machine has reported a delayed heartbeat to the central monitoring service.",
        "component": "Network Interface",
        "severity": "Sev4",
        "service": "Monitoring remains functional, but connectivity quality is degraded and should be checked.",
        "technician": "Inspect the network path, reboot the communication module, and confirm stable connectivity after restart.",
    },
    {
        "title": "Cup Sensor State Mismatch",
        "description": "The cup detection sensor is intermittently miscounting the cup tray state.",
        "component": "Cup Dispenser",
        "severity": "Sev4",
        "service": "The machine remains functional but may miscount cups during busy periods.",
        "technician": "Realign the sensor and verify the dispenser reports the correct cup count after a full cycle.",
    },
    {
        "title": "Maintenance Reminder Still Pending",
        "description": "A scheduled preventive action has not yet been acknowledged by the maintenance team.",
        "component": "Service Scheduler",
        "severity": "Sev3",
        "service": "The machine is operational but preventive maintenance remains due.",
        "technician": "Complete the outstanding maintenance task and clear the reminder from the service dashboard.",
    },
    {
        "title": "Dispense Time Overrun Warning",
        "description": "A recent beverage cycle exceeded the normal dispense time by more than eight seconds.",
        "component": "Pump Assembly",
        "severity": "Sev3",
        "service": "The beverage still dispenses, but timing is outside the expected range and should be corrected.",
        "technician": "Clean the nozzle and verify pump flow rate and valve timing before resuming normal operation.",
    },
    {
        "title": "Card Reader Checksum Error",
        "description": "The payment reader reported intermittent checksum errors during transaction processing.",
        "component": "Payment Reader Board",
        "severity": "Sev3",
        "service": "Transaction completion is possible but unstable and may require intervention during busy periods.",
        "technician": "Reseat the connector, inspect the reader board, and validate transaction integrity across test cards.",
    },
    {
        "title": "Ingredient Mix Ratio Drift",
        "description": "The machine detected an out-of-spec ingredient mix during a brew cycle.",
        "component": "Brew Group",
        "severity": "Sev3",
        "service": "The drink is still produced but recipe consistency is degraded and should be corrected soon.",
        "technician": "Flush the brew group and recalibrate the ingredient dose timing and solenoid actuation.",
    },
    {
        "title": "Boiler Pressure Trending High",
        "description": "Boiler pressure is above its preferred range for several recent cycles.",
        "component": "Boiler Control",
        "severity": "Sev4",
        "service": "The machine remains available, but pressure control should be inspected on the next service visit.",
        "technician": "Inspect the regulator and verify the pressure switch and safety valve are responding correctly.",
    },
    {
        "title": "Cabinet Fan Cycling High",
        "description": "Ventilation fan activity has exceeded the expected frequency for this machine location.",
        "component": "Cooling Fan",
        "severity": "Sev4",
        "service": "Cooling remains active, but the fan cycle trend indicates a need for maintenance review.",
        "technician": "Check for airflow restrictions and lubrication issues before the next peak demand period.",
    },
]

critical_templates = [
    {
        "title": "Payment Transactions Blocked Now",
        "description": "The cashless payment system is unable to authorize any transactions and the machine is blocked.",
        "component": "Payment Terminal",
        "severity": "Sev1",
        "service": "The machine cannot accept cashless payments and must be repaired before service resumes.",
        "technician": "Inspect the card reader hardware, verify all wiring, and replace the terminal if the self-test fails.",
    },
    {
        "title": "Boiler Temperature Overheat Detected",
        "description": "The boiler has exceeded the safe operating temperature and entered a protective shutdown state.",
        "component": "Boiler and Heating Circuit",
        "severity": "Sev1",
        "service": "Hot beverage service is suspended until the thermal fault is cleared and the heater control is validated.",
        "technician": "Power down the unit, verify thermostat and thermal fuse operation, and replace failed heating components.",
    },
    {
        "title": "Water Supply Connection Lost",
        "description": "The machine cannot complete beverage preparation because the water supply is unavailable.",
        "component": "Water Reservoir",
        "severity": "Sev2",
        "service": "Beverage service is unavailable until the system is refilled and the inlet path is verified.",
        "technician": "Refill the reservoir, inspect the inlet valve, and confirm the pump starts normally under load.",
    },
    {
        "title": "Main Controller Hard Fault",
        "description": "The main controller failed to complete startup initialization and entered a hard fault state.",
        "component": "Main Controller",
        "severity": "Sev1",
        "service": "The unit cannot serve customers until the controller has been restarted and verified.",
        "technician": "Reboot the controller, reflash firmware if needed, and validate all sensor states before returning service.",
    },
    {
        "title": "Coffee Grinder Mechanism Jammed",
        "description": "The grinder is blocked and all beverage dispensing operations are stopped.",
        "component": "Coffee Grinder",
        "severity": "Sev2",
        "service": "No beverage preparation can continue until the grinder is unclogged and tested.",
        "technician": "Clear the jam, inspect the grind chamber, and verify the assembly runs normally through a calibration cycle.",
    },
    {
        "title": "Boiler Leak Now Detected",
        "description": "A critical leak was detected in the boiler chamber and the machine was safely shut down.",
        "component": "Boiler Chamber",
        "severity": "Sev1",
        "service": "The machine must remain offline until the chamber and seals have been inspected and repaired.",
        "technician": "Isolate the water supply, inspect the chamber seal and body, and replace any failed components before restart.",
    },
    {
        "title": "Milk Dispense Pump Failed",
        "description": "The milk pumping circuit has failed and dairy-based beverages are unavailable.",
        "component": "Milk Pump",
        "severity": "Sev2",
        "service": "Milk beverages cannot be served until the pump is replaced and the circuit is primed.",
        "technician": "Replace the pump assembly, prime the line, and validate flow across three test cycles.",
    },
    {
        "title": "Main Power Supply Failure",
        "description": "The primary power supply is unstable and is interrupting machine operation.",
        "component": "Power Supply Unit",
        "severity": "Sev1",
        "service": "The machine has dropped functions and must remain offline until the PSU is replaced.",
        "technician": "Measure output voltage under load, replace the failing supply, and verify no downstream board is experiencing brown-out conditions.",
    },
    {
        "title": "Safety Door Lockout Engaged",
        "description": "A door safety lockout has engaged, preventing all service operations from resuming.",
        "component": "Door Interlock",
        "severity": "Sev1",
        "service": "The machine is locked and cannot operate until the latch state is corrected and reset.",
        "technician": "Inspect the actuator, verify the sensor wiring, and clear the lockout before restoring service.",
    },
    {
        "title": "Dispense Pump Cavitation Detected",
        "description": "The dispenser pump is cavitating and no longer delivers a reliable beverage flow.",
        "component": "Dispense Pump",
        "severity": "Sev1",
        "service": "Beverage service is stopped until the pump is primed or replaced and tested.",
        "technician": "Prime the pump, inspect for air ingress or blockage, and replace the pump if pressure recovery is not achieved.",
    },
]

alerts = []


def validate_alert_title(title, minimum_words=3, maximum_words=5):
    """Keep generated titles within the compact range used by machine alerts."""
    word_count = len(re.findall(r"\b[\w'-]+\b", title))
    if not minimum_words <= word_count <= maximum_words:
        raise ValueError(
            f"Alert title must contain {minimum_words} to {maximum_words} words: {title!r} "
            f"contains {word_count}."
        )


for template in info_templates + warning_templates + critical_templates:
    validate_alert_title(template["title"])

for idx in range(19):
    template = info_templates[idx % len(info_templates)]
    system_area, group = system_area_taxonomy(template["component"])
    alerts.append({
        "type": "Informational",
        "severity": "Sev5",
        "title": template["title"],
        "description": template["description"],
        "system_area": system_area,
        "group": group,
        "service": template["service"],
        "technician": template["technician"],
    })
for idx in range(40):
    template = warning_templates[idx % len(warning_templates)]
    system_area, group = system_area_taxonomy(template["component"])
    alerts.append({
        "type": "Warning",
        "severity": template["severity"],
        "title": template["title"],
        "description": template["description"],
        "system_area": system_area,
        "group": group,
        "service": template["service"],
        "technician": template["technician"],
    })
for idx in range(16):
    template = critical_templates[idx % len(critical_templates)]
    system_area, group = system_area_taxonomy(template["component"])
    alerts.append({
        "type": "Critical",
        "severity": template["severity"],
        "title": template["title"],
        "description": template["description"],
        "system_area": system_area,
        "group": group,
        "service": template["service"],
        "technician": template["technician"],
    })

headers = [
    "ID",
    "Alert Title",
    "Type",
    "Severity",
    "Alert Description",
    "System Area",
    "Major Group",
    "Operator Response",
    "Model",
    "Last Update",
    "Version",
    "Notes",
    "Critical Stop Response",
    "Service Response",
    "Technician Response",
]

operator_responses = {
    "Maintenance Reminder Still Pending": "Run service restart.\nRun maintenance cycle.",
    "Bean Hopper Level Low": "Refill bean hopper.",
    "Milk Tank Below Refill Target": "Refill milk tank.",
    "Water Tank Level Warning": "Refill water tank.",
    "Payment Retry Rate Elevated": "Disconnect and reconnect payment machine.\nTry with secondary payment method.\nReconnect primary machine.",
    "Drip Tray Nearly Full": "Empty drip tray.",
    "Payment Transactions Blocked Now": "Check network is up.\nFollow payment machine restart procedure.\nRetry again.\nIf problem persists, call service team.",
    "Boiler Temperature Overheat Detected": "Wait for 15 minutes.\nTry again.",
    "Main Controller Hard Fault": "Follow machine restart procedure.\nIf problem persists, call service team.",
    "Coffee Grinder Mechanism Jammed": "Empty coffee grinder.\nClean hopper.\nRestart machine.\nTry again.\nIf issue persists, call service team.",
    "Boiler Leak Now Detected": "Turn off machine.\nEmpty tank.\nCall service team.",
    "Main Power Supply Failure": "Do not use machine.\nTurn off mains power.\nCall service team.",
    "Safety Door Lockout Engaged": "Check door latch.\nCheck for occlusions.\nRestart machine.",
    "Dispense Pump Cavitation Detected": "Turn off machine.\nRun coffee maintenance cycle twice.\nRetry.\nIf problem persists, call service team.",
}

OPERATOR_VERIFICATION_CHECKS = {
    "Control & Interface": (
        "Confirm the display and status indicators match the alert condition.",
        "Retry the affected operation once and note the result.",
        "Record any visible error code before escalating the issue.",
    ),
    "Beverage Systems": (
        "Check for visible leaks, blockages, or unusual noise.",
        "Run one safe test cycle and confirm the expected flow.",
        "Keep the affected beverage unavailable if the check fails.",
    ),
    "Supply & Payment": (
        "Confirm stock levels or payment connectivity at the machine.",
        "Perform one test transaction or dispense check where safe.",
        "Record the outcome and escalate if normal operation is not restored.",
    ),
    "Operations & Safety": (
        "Inspect the surrounding area for an immediate safety concern.",
        "Confirm doors, trays, vents, and access panels are correctly seated.",
        "Keep the machine out of service if the alert remains active.",
    ),
}

SERVICE_VERIFICATION_CHECKS = {
    "Control & Interface": (
        "Review the event log and confirm the alert timestamp.",
        "Verify communication, display, and controller health checks pass.",
        "Document the result before closing or escalating the service case.",
    ),
    "Beverage Systems": (
        "Inspect the relevant water, heating, and dispense path.",
        "Complete a controlled test cycle and compare readings with specification.",
        "Confirm there are no leaks or recurring faults before return to service.",
    ),
    "Supply & Payment": (
        "Verify supply levels, sensors, connections, and peripheral status.",
        "Complete a test transaction or product cycle as applicable.",
        "Confirm counters and monitoring data update correctly after the test.",
    ),
    "Operations & Safety": (
        "Inspect safety interlocks, ventilation, power, and service access points.",
        "Clear the alert only after the relevant verification check passes.",
        "Record corrective action and confirm the machine is safe to operate.",
    ),
}


def bullet_list(primary_response, additional_items, default_response):
    """Format an existing response and three supporting checks as bullets."""
    primary_items = [line.strip() for line in primary_response.splitlines() if line.strip()]
    if not primary_items:
        primary_items = [default_response]
    return "\n".join(f"• {item}" for item in (*primary_items, *additional_items))

models = ("All", "SC1x, SC3x", "SC2x", "SC2x, SC3x")
column_widths = {
    "A": 8.16, "B": 31, "C": 18, "E": 41.66, "F": 13.66,
    "G": 21, "H": 27.66, "I": 13, "J": 14.16, "L": 17.33,
    "M": 18, "N": 32.66, "O": 29.66,
}

HUMAN_ERROR_REPLACEMENTS = {
    "maintenance": "maintenece",
    "hopper": "hopr",
    "transaction": "txn",
    "printer": "prntr",
    "paper": "ppr",
    "machine": "maschine",
}


def add_misspellings(text, rng, error_rate=0.25):
    """Add occasional configured misspellings to a string."""
    if not text:
        return text

    result = text
    for correct, typo in HUMAN_ERROR_REPLACEMENTS.items():
        pattern = rf"\b{correct}\b"
        if re.search(pattern, result, flags=re.IGNORECASE) and rng.random() < error_rate:
            result = re.sub(
                pattern,
                lambda match: typo.capitalize() if match.group(0)[0].isupper() else typo,
                result,
                flags=re.IGNORECASE,
            )
    return result


def add_human_errors(text, rng, error_rate=0.25):
    """Add occasional misspellings and punctuation errors to a response string."""
    if not text:
        return text

    result = add_misspellings(text, rng, error_rate)

    if rng.random() < error_rate:
        result = re.sub(r"\. (?=\S)", ".  ", result)
        result = result.replace(".\n", ".  \n")
    if result.endswith(".") and rng.random() < error_rate:
        result = result[:-1]
    return result


def _row_height(values):
    """Estimate Excel's wrapped row height while allowing multi-step responses."""
    widths = [column_widths.get(get_column_letter(i), 8.83) for i in range(1, len(headers) + 1)]
    line_count = 1
    for value, width in zip(values, widths):
        text = "" if value is None else str(value)
        wrapped = sum(max(1, (len(line) + max(1, int(width)) - 1) // max(1, int(width))) for line in text.split("\n"))
        line_count = max(line_count, wrapped)
    return min(180, max(48, line_count * 16))


def build_workbook(human_errors=False, human_error_rate=0.25, human_error_seed=21):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Alerts"
    worksheet.append(headers)
    error_rng = random.Random(human_error_seed)

    first_update = date(2023, 12, 15)
    for index, alert in enumerate(alerts, start=1):
        last_update = first_update + timedelta(days=index - 1)
        version_year = 2025 if alert["type"] == "Critical" else 2023
        description = alert["description"]
        operator_response = bullet_list(
            operator_responses.get(alert["title"], ""),
            OPERATOR_VERIFICATION_CHECKS[alert["group"]],
            "No immediate operator action is required; continue monitoring the machine.",
        )
        service_response = bullet_list(
            alert["service"],
            SERVICE_VERIFICATION_CHECKS[alert["group"]],
            "Review the alert and confirm the machine state.",
        )
        technician_response = alert["technician"]
        if human_errors:
            description = add_misspellings(description, error_rng, human_error_rate)
            operator_response = add_human_errors(operator_response, error_rng, human_error_rate)
            service_response = add_human_errors(service_response, error_rng, human_error_rate)
            technician_response = add_human_errors(technician_response, error_rng, human_error_rate)

        row = [
            str(index),
            alert["title"],
            alert["type"],
            alert["severity"],
            description,
            alert["system_area"],
            alert["group"],
            operator_response,
            models[(index - 1) // 19 % len(models)],
            last_update,
            f"DT 24.{version_year}.{index:02d}",
            "",
            "Yes" if alert["type"] == "Critical" else "",
            service_response,
            technician_response,
        ]
        worksheet.append(row)
        worksheet.row_dimensions[worksheet.max_row].height = _row_height(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet["A1"].number_format = "@"
    for cell in worksheet["A"][1:]:
        cell.number_format = "@"
    for cell in worksheet["J"][1:]:
        cell.number_format = "d-mmm-yy"

    worksheet.freeze_panes = "A2"
    last_column = get_column_letter(len(headers))
    worksheet.auto_filter.ref = f"A1:{last_column}{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 150
    return workbook


def main():
    parser = argparse.ArgumentParser(description="Generate the cashless machine alert catalogue.")
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT, help="Output .xlsx path")
    parser.add_argument(
        "--human-errors",
        action="store_true",
        help="Add occasional intentional errors to response columns only",
    )
    parser.add_argument(
        "--human-error-rate",
        type=float,
        default=0.25,
        metavar="RATE",
        help="Probability of each possible response error (0.0 to 1.0; default: 0.25)",
    )
    parser.add_argument(
        "--human-error-seed",
        type=int,
        default=21,
        metavar="SEED",
        help="Random seed for reproducible intentional errors (default: 21)",
    )
    args = parser.parse_args()
    output = args.output
    if "donotdelete" in output.name.lower():
        parser.error("refusing to overwrite a DONOTDELETE reference workbook")
    if output.suffix.lower() != ".xlsx":
        parser.error("output must use the .xlsx extension")
    if not 0.0 <= args.human_error_rate <= 1.0:
        parser.error("--human-error-rate must be between 0.0 and 1.0")
    output.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(
        human_errors=args.human_errors,
        human_error_rate=args.human_error_rate,
        human_error_seed=args.human_error_seed,
    ).save(output)
    mode = " with intentional human errors" if args.human_errors else ""
    print(f"Created {output} with {len(alerts)} alerts{mode}.")


if __name__ == "__main__":
    main()
