import argparse
import json
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = PROJECT_ROOT / "data" / "source" / "alerts-ds.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "alerts-ds.json"


def json_value(value: Any) -> Any:
    """Convert Excel values that JSON cannot serialize natively."""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == time.min else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def extract_rows(workbook_path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name is not None and sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(f"sheet {sheet_name!r} not found; available sheets: {available}")

        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = worksheet.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration as error:
            raise ValueError(f"sheet {worksheet.title!r} is empty") from error

        headers = [str(value).strip() if value is not None else "" for value in header_row]
        if any(not header for header in headers):
            raise ValueError(f"sheet {worksheet.title!r} contains a blank header")
        if len(headers) != len(set(headers)):
            raise ValueError(f"sheet {worksheet.title!r} contains duplicate headers")

        records = []
        for row in rows:
            values = list(row[: len(headers)])
            values.extend([None] * (len(headers) - len(values)))
            if all(value is None for value in values):
                continue
            records.append(dict(zip(headers, map(json_value, values))))
        return records
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract an Excel alert sheet to JSON.")
    parser.add_argument("input", nargs="?", default=DEFAULT_INPUT, type=Path)
    parser.add_argument("-o", "--output", type=Path, help="Defaults to the input filename with a .json extension")
    parser.add_argument("--sheet", help="Sheet to extract; defaults to the active sheet")
    args = parser.parse_args()

    if not args.input.is_file():
        parser.error(f"input workbook not found: {args.input}")

    output = args.output or DEFAULT_OUTPUT
    output.parent.mkdir(parents=True, exist_ok=True)

    try:
        records = extract_rows(args.input, args.sheet)
    except ValueError as error:
        parser.error(str(error))

    source_modified_at = datetime.fromtimestamp(args.input.stat().st_mtime, timezone.utc)
    payload = {
        "metadata": {
            "schema_version": 1,
            "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "source_file": args.input.name,
            "source_modified_at": source_modified_at.isoformat(timespec="seconds"),
            "record_count": len(records),
        },
        "alerts": records,
    }

    with output.open("w", encoding="utf-8") as json_file:
        json.dump(payload, json_file, ensure_ascii=False, indent=2)
        json_file.write("\n")

    print(f"Extracted {len(records)} records from {args.input} to {output}.")


if __name__ == "__main__":
    main()
