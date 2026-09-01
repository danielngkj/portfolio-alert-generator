import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook, Workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = PROJECT_ROOT / "data" / "source" / "alerts-ds.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "generated" / "alerts-ds-clean.xlsx"
DEFAULT_REPORT = PROJECT_ROOT / "data" / "quality" / "language-cleanup-report.json"

REPLACEMENTS = {
    "maintenece": "maintenance",
    "hopr": "hopper",
    "txn": "transaction",
    "prntr": "printer",
    "ppr": "paper",
    "maschine": "machine",
    "maintance": "maintenance",
}

STANDARD_ABBREVIATIONS = {
    "ID": "ID",
    "UI": "UI",
    "UPS": "UPS",
    "PDU": "PDU",
    "HMI": "HMI",
    "PLC": "PLC",
    "LCD": "LCD",
    "LED": "LED",
    "RPM": "RPM",
    "VFD": "VFD",
    "DC": "DC",
    "AC": "AC",
    "PSU": "PSU",
    "PCB": "PCB",
    "N/A": "N/A",
    "S/N": "S/N",
}

TEXT_FIELDS = [
    "Alert Description",
    "Operator Response",
    "Service Response",
    "Technician Response",
]


def normalize_text(text: str) -> str:
    if text is None:
        return ""
    cleaned = str(text).strip()
    cleaned = re.sub(r"\.\s{2,}", ". ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned)
    if cleaned and cleaned[-1] not in ".!?":
        cleaned = cleaned + "."
    return cleaned


def cleanup_text(text: str) -> str:
    if text is None:
        return ""
    cleaned = str(text)
    for wrong, correct in REPLACEMENTS.items():
        pattern = re.compile(rf"\b{re.escape(wrong)}\b", flags=re.IGNORECASE)
        cleaned = pattern.sub(lambda match: correct if match.group(0)[0].islower() else correct.capitalize(), cleaned)
    lines = cleaned.split("\n")
    cleaned_lines = []
    for line in lines:
        line = re.sub(r"\.\s{2,}", ". ", line)
        line = re.sub(r"[ \t]+", " ", line)
        line = line.replace("..", ".")
        line = line.strip()
        if line and line[-1] not in ".!?":
            line += "."
        cleaned_lines.append(line)
    cleaned = "\n".join(cleaned_lines)
    return cleaned


def get_row_values(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return [], []
        headers = list(rows[0])
        data_rows = rows[1:]
        return headers, data_rows
    finally:
        workbook.close()


def build_clean_workbook(source_path: Path, output_path: Path):
    headers, rows = get_row_values(source_path)
    cleaned_workbook = Workbook()
    cleaned_sheet = cleaned_workbook.active
    cleaned_sheet.title = "Alerts"
    cleaned_sheet.append(headers)

    for row in rows:
        cleaned_row = list(row)
        for index, value in enumerate(cleaned_row):
            header = headers[index] if index < len(headers) else ""
            if header in TEXT_FIELDS and value is not None:
                cleaned_row[index] = cleanup_text(value)
        cleaned_sheet.append(cleaned_row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    cleaned_workbook.save(output_path)
    return output_path


def build_report(source_path: Path, output_path: Path):
    headers, rows = get_row_values(source_path)
    issues = []
    fields = TEXT_FIELDS
    for row_index, row in enumerate(rows, start=2):
        record = dict(zip(headers, row))
        for field in fields:
            value = record.get(field)
            if value is None:
                continue
            original = str(value)
            cleaned = cleanup_text(original)
            if cleaned != original:
                issues.append({
                    "row": row_index,
                    "field": field,
                    "before": original,
                    "after": cleaned,
                })
    report = {
        "source_file": str(source_path),
        "output_file": str(output_path),
        "total_corrections": len(issues),
        "examples": issues[:10],
    }
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Clean language-level issues in the generated alert workbook.")
    parser.add_argument("input", nargs="?", default=DEFAULT_INPUT, type=Path, help="Workbook to clean")
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT, help="Cleaned workbook destination")
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT, help="JSON report of cleanup actions")
    args = parser.parse_args()

    if not args.input.is_file():
        parser.error(f"input workbook not found: {args.input}")

    output_path = build_clean_workbook(args.input, args.output)
    report = build_report(args.input, output_path)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    with args.report.open("w", encoding="utf-8") as report_file:
        json.dump(report, report_file, ensure_ascii=False, indent=2)
        report_file.write("\n")

    print(f"Created cleaned workbook at {output_path}.")
    print(f"Cleanup summary: {report['total_corrections']} text changes recorded.")


if __name__ == "__main__":
    main()
