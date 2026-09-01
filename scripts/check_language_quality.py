import argparse
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = PROJECT_ROOT / "data" / "source" / "alerts-ds.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "quality" / "alert-language-quality-report.json"

MISSPELLINGS = {
    "maintenance": "maintenance",
    "maintenece": "maintenance",
    "hopper": "hopper",
    "hopr": "hopper",
    "transaction": "transaction",
    "txn": "transaction",
    "printer": "printer",
    "prntr": "printer",
    "paper": "paper",
    "ppr": "paper",
    "machine": "machine",
    "maschine": "machine",
}

ABBREVIATION_TERMS = {
    "ID",
    "UI",
    "UPS",
    "PDU",
    "HMI",
    "PLC",
    "LCD",
    "LED",
    "RPM",
    "VFD",
    "DC",
    "AC",
    "PSU",
    "PCB",
    "N/A",
    "S/N",
}

TEXT_FIELDS = [
    "Alert Description",
    "Operator Response",
    "Service Response",
    "Technician Response",
]


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\.\s{2,}", ". ", text)
    text = re.sub(r"\s+", " ", text)
    return text


def find_misspellings(text):
    matches = []
    for wrong, correct in MISSPELLINGS.items():
        pattern = re.compile(rf"\b{re.escape(wrong)}\b", flags=re.IGNORECASE)
        if pattern.search(text):
            matches.append({"wrong": wrong, "correct": correct})
    return matches


def find_abbreviation_issues(text):
    matches = []
    for token in re.findall(r"\b[A-Z]{2,5}\b", text):
        if token in ABBREVIATION_TERMS:
            matches.append(token)
    return matches


def field_issues(value, field_name):
    text = normalize_text(value)
    issues = []

    if not text:
        return issues

    if re.search(r"\.\s{2,}", str(value or "")):
        issues.append({"rule": "double_space_after_period", "detail": "Double spacing after a period was detected."})

    if field_name in {"Operator Response", "Service Response", "Technician Response"}:
        if not text.endswith((".", "!", "?")):
            issues.append({"rule": "missing_terminal_punctuation", "detail": "Response text should end with terminal punctuation."})

    for miss in find_misspellings(text):
        issues.append({"rule": "misspelling", "detail": f"Possible misspelling: {miss['wrong']} -> {miss['correct']}"})

    abbreviations = find_abbreviation_issues(text)
    if abbreviations:
        issues.append({"rule": "abbreviation_usage", "detail": f"Abbreviation usage detected: {', '.join(sorted(set(abbreviations)))}"})

    return issues


def collect_quality_rows(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        field_positions = {name: index for index, name in enumerate(headers)}
        records = []

        for row_number, row in enumerate(rows[1:], start=2):
            if row is None or all(value is None for value in row):
                continue
            record = dict(zip(headers, row))
            alert_id = record.get("ID") or record.get("id") or row_number
            findings = []
            for field_name in TEXT_FIELDS:
                if field_name not in field_positions:
                    continue
                value = record.get(field_name)
                for issue in field_issues(value, field_name):
                    findings.append({
                        "alert_id": str(alert_id),
                        "field": field_name,
                        "rule": issue["rule"],
                        "detail": issue["detail"],
                        "row": row_number,
                    })
            if findings:
                records.extend(findings)
        return records
    finally:
        workbook.close()


def build_report(workbook_path: Path):
    findings = collect_quality_rows(workbook_path)
    counts = Counter(item["rule"] for item in findings)
    report = {
        "metadata": {
            "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "source_file": workbook_path.name,
            "source_path": str(workbook_path),
            "fields_checked": TEXT_FIELDS,
            "total_findings": len(findings),
        },
        "summary": {
            "rule_counts": dict(sorted(counts.items())),
            "issues_by_field": dict(sorted(Counter(item["field"] for item in findings).items())),
        },
        "findings": findings[:25],
    }
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Check source text quality for generated alert data.")
    parser.add_argument("input", nargs="?", default=DEFAULT_INPUT, type=Path, help="Workbook to scan")
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT, help="Path to JSON quality report")
    args = parser.parse_args()

    if not args.input.is_file():
        parser.error(f"input workbook not found: {args.input}")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    report = build_report(args.input)
    with args.output.open("w", encoding="utf-8") as report_file:
        json.dump(report, report_file, ensure_ascii=False, indent=2)
        report_file.write("\n")
    print(f"Scanned {args.input} and wrote quality report to {args.output}.")
    print(json.dumps(report["summary"]["rule_counts"], indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
