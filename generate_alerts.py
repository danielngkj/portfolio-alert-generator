import argparse
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

taxonomy_map = {
    "Control Board": "Control",
    "Connectivity Module": "Control",
    "Cleaning System": "Maintenance",
    "Ingredient Hoppers": "Ingredients",
    "Main Controller": "Control",
    "Payment Terminal": "Payment",
    "Boiler and Tank Assembly": "Brew",
    "Water Supply Line": "Water",
    "Dispense Nozzle": "Brew",
    "Door Interlock": "Safety",
    "Payment Gateway Interface": "Payment",
    "Receipt Printer": "Display",
    "Bean Hopper": "Ingredients",
    "Milk Reservoir": "Ingredients",
    "Water Tank": "Water",
    "Card Reader": "Payment",
    "Touchscreen Display": "Display",
    "Drip Tray": "Maintenance",
    "Pump Motor": "Brew",
    "Dispense Valve": "Brew",
    "Network Interface": "Control",
    "Cup Dispenser": "Dispense",
    "Service Scheduler": "Maintenance",
    "Pump Assembly": "Brew",
    "Payment Reader Board": "Payment",
    "Brew Group": "Brew",
    "Boiler Control": "Brew",
    "Cooling Fan": "Cooling",
    "Boiler Chamber": "Brew",
    "Milk Pump": "Ingredients",
    "Power Supply Unit": "Power",
    "Coffee Grinder": "Brew",
    "Water Reservoir": "Water",
    "Boiler and Heating Circuit": "Brew",
    "Customer Interface": "Display",
    "Electrical System": "Power",
    "Thermal Management": "Cooling",
    "Safety & Access": "Safety",
    "Beverage Handling": "Dispense",
    "Ingredient Handling": "Ingredients",
    "Water System": "Water",
    "Beverage Preparation": "Brew",
    "Machine Control": "Control",
    "Maintenance": "Maintenance",
    "Cashless Payment": "Payment",
    "Display": "Display",
    "Cooling": "Cooling",
    "Power": "Power",
    "Safety": "Safety",
    "Dispense": "Dispense",
    "Ingredients": "Ingredients",
    "Water": "Water",
    "Brew": "Brew",
}


info_templates = [
    {
        "title": "Startup Complete",
        "description": "Machine completed a normal startup sequence without faults.",
        "component": "Control Board",
        "service": "No service action required; the machine is operating within normal parameters.",
        "technician": "Log the event and confirm normal startup values remain stable across the next operating cycle.",
    },
    {
        "title": "Heartbeat Confirmed",
        "description": "The device is successfully reporting to the service backend.",
        "component": "Connectivity Module",
        "service": "Remote monitoring connection is healthy and no intervention is needed.",
        "technician": "Review heartbeat timestamps and confirm they are within the expected interval for this site.",
    },
    {
        "title": "Scheduled Cleaning Logged",
        "description": "The maintenance cleaning cycle completed without interruption.",
        "component": "Cleaning System",
        "service": "Cleaning cycle completed successfully and the unit remained in service.",
        "technician": "Verify the cleaning log and confirm the next maintenance cycle is scheduled correctly.",
    },
    {
        "title": "Inventory Balanced",
        "description": "Ingredient and product inventory status is within expected limits.",
        "component": "Ingredient Hoppers",
        "service": "Inventory counts are aligned with the machine configuration and no refill is needed.",
        "technician": "Cross-check inventory counters against the physical stock levels for reporting accuracy.",
    },
    {
        "title": "Firmware Check Passed",
        "description": "Current firmware version matches the approved release.",
        "component": "Main Controller",
        "service": "No patch or rollback is required at this time.",
        "technician": "Archive the firmware record and validate version alignment against the approved release list.",
    },
    {
        "title": "Payment Self-Test OK",
        "description": "The cashless payment reader completed its self-diagnostic routine successfully.",
        "component": "Payment Terminal",
        "service": "Cashless device is fit for customer transactions and no service response is needed.",
        "technician": "Retain the diagnostic record and monitor for any future reader instability.",
    },
    {
        "title": "Tank Temperature Stable",
        "description": "Beverage tank temperature remains within the defined operating band.",
        "component": "Boiler and Tank Assembly",
        "service": "Temperature regulation is stable and no user-facing issue is present.",
        "technician": "Check the last thermal trend and confirm no drift is emerging from previous service activity.",
    },
    {
        "title": "Water Quality Normal",
        "description": "Water quality readings are within the expected operational range.",
        "component": "Water Supply Line",
        "service": "Water conditions are acceptable and no service escalation is necessary.",
        "technician": "Note the water quality trend for reference during the next preventive inspection.",
    },
    {
        "title": "Nozzle Purge Successful",
        "description": "The dispense nozzle purge cycle completed without blockage or abnormal flow.",
        "component": "Dispense Nozzle",
        "service": "The purge cycle completed successfully and the machine remains ready for service.",
        "technician": "Log the nozzle purge outcome and confirm flow metrics remain within expected values.",
    },
    {
        "title": "Door Sensor Normal",
        "description": "Door interlock and access sensors are in the normal state.",
        "component": "Door Interlock",
        "service": "No security or access event is active and the machine is operating as expected.",
        "technician": "Monitor sensor state during the next operating cycle to confirm the latch remains stable.",
    },
    {
        "title": "Transaction Completed",
        "description": "A recent customer transaction ended normally with no duplicate settlement event.",
        "component": "Payment Gateway Interface",
        "service": "The transaction lifecycle completed correctly with no response anomaly.",
        "technician": "Confirm settlement logs match the backend report and archive the transaction record.",
    },
    {
        "title": "Receipt Queue Cleared",
        "description": "Receipt printing queue is empty and journal activity is normal.",
        "component": "Receipt Printer",
        "service": "Printer queue was cleared successfully and no paper or jam issue is present.",
        "technician": "Inspect remaining paper stock and confirm print quality remains acceptable.",
    },
]

warning_templates = [
    {
        "title": "Low Bean Hopper",
        "description": "The coffee bean hopper level is approaching the minimum threshold for continuous service.",
        "component": "Bean Hopper",
        "severity": "Sev4",
        "service": "Refill is recommended before the next customer peak period to avoid service disruption.",
        "technician": "Top up the hopper and confirm the level sensor reads correctly after refill.",
    },
    {
        "title": "Milk Tank Below Target",
        "description": "Milk inventory is beneath the preferred refill limit but the machine remains operational.",
        "component": "Milk Reservoir",
        "severity": "Sev4",
        "service": "The unit is still serving beverages, but milk supply should be replenished soon.",
        "technician": "Refill the container and recalibrate the float sensor if the reading remains offset.",
    },
    {
        "title": "Water Level Warning",
        "description": "The water tank is nearing the low-level threshold and may require replenishment soon.",
        "component": "Water Tank",
        "severity": "Sev4",
        "service": "The machine can continue operating, but refill should be scheduled before the next service window.",
        "technician": "Fill the tank and inspect the level sensor for a delayed or sticky float response.",
    },
    {
        "title": "Payment Retry Rate Elevated",
        "description": "The payment system is showing a slightly higher-than-normal retry rate.",
        "component": "Card Reader",
        "severity": "Sev4",
        "service": "Transactions are still completing, but card acceptance should be monitored for intermittent faults.",
        "technician": "Clean the reader slot and verify contactless and contact reads are consistently accepted.",
    },
    {
        "title": "Touchscreen Response Slow",
        "description": "The HMI touch panel is responding slower than the expected baseline.",
        "component": "Touchscreen Display",
        "severity": "Sev4",
        "service": "Customer interaction remains possible, but responsiveness is degrading and should be corrected.",
        "technician": "Clean the panel and recalibrate the touch matrix to restore expected response speed.",
    },
    {
        "title": "Drip Tray Almost Full",
        "description": "The drip tray sensor reports the tray level approaching the service threshold.",
        "component": "Drip Tray",
        "severity": "Sev4",
        "service": "The machine remains available, but the tray should be emptied before the next operating cycle.",
        "technician": "Empty the tray, reset the sensor count, and verify the tray is reporting empty correctly.",
    },
    {
        "title": "Pump Delay Notice",
        "description": "A soft start delay was observed before beverage dispensing began.",
        "component": "Pump Motor",
        "severity": "Sev4",
        "service": "Dispense still works, but delivery timing is slower than expected and should be reviewed.",
        "technician": "Inspect the pump startup sequence and validate the control board is not limiting power unnecessarily.",
    },
    {
        "title": "Recipe Volume Low",
        "description": "One beverage recipe is dispensing slightly below the configured target volume.",
        "component": "Dispense Valve",
        "severity": "Sev4",
        "service": "Service is still available, but the recipe volume should be recalibrated during the next visit.",
        "technician": "Recalibrate the recipe volume and test three consecutive servings before returning the unit.",
    },
    {
        "title": "Delayed Heartbeat",
        "description": "The machine has reported a delayed heartbeat to the central monitoring service.",
        "component": "Network Interface",
        "severity": "Sev4",
        "service": "Monitoring remains functional, but connectivity quality is degraded and should be checked.",
        "technician": "Inspect the network path, reboot the communication module, and confirm stable connectivity after restart.",
    },
    {
        "title": "Cup Sensor Mismatch",
        "description": "The cup detection sensor is intermittently miscounting the cup tray state.",
        "component": "Cup Dispenser",
        "severity": "Sev4",
        "service": "The machine remains functional but may miscount cups during busy periods.",
        "technician": "Realign the sensor and verify the dispenser reports the correct cup count after a full cycle.",
    },
    {
        "title": "Maintenance Reminder Pending",
        "description": "A scheduled preventive action has not yet been acknowledged by the maintenance team.",
        "component": "Service Scheduler",
        "severity": "Sev3",
        "service": "The machine is operational but preventive maintenance remains due.",
        "technician": "Complete the outstanding maintenance task and clear the reminder from the service dashboard.",
    },
    {
        "title": "Dispense Time Overrun",
        "description": "A recent beverage cycle exceeded the normal dispense time by more than eight seconds.",
        "component": "Pump Assembly",
        "severity": "Sev3",
        "service": "The beverage still dispenses, but timing is outside the expected range and should be corrected.",
        "technician": "Clean the nozzle and verify pump flow rate and valve timing before resuming normal operation.",
    },
    {
        "title": "Card Checksum Error",
        "description": "The payment reader reported intermittent checksum errors during transaction processing.",
        "component": "Payment Reader Board",
        "severity": "Sev3",
        "service": "Transaction completion is possible but unstable and may require intervention during busy periods.",
        "technician": "Reseat the connector, inspect the reader board, and validate transaction integrity across test cards.",
    },
    {
        "title": "Ingredient Ratio Drift",
        "description": "The machine detected an out-of-spec ingredient mix during a brew cycle.",
        "component": "Brew Group",
        "severity": "Sev3",
        "service": "The drink is still produced but recipe consistency is degraded and should be corrected soon.",
        "technician": "Flush the brew group and recalibrate the ingredient dose timing and solenoid actuation.",
    },
    {
        "title": "Boiler Pressure Trending High",
        "description": "Boiler pressure is above its preferred range for several recent cycles.",
        "component": "Boiler Control",
        "severity": "Sev4",
        "service": "The machine remains available, but pressure control should be inspected on the next service visit.",
        "technician": "Inspect the regulator and verify the pressure switch and safety valve are responding correctly.",
    },
    {
        "title": "Cabinet Fan Cycle High",
        "description": "Ventilation fan activity has exceeded the expected frequency for this machine location.",
        "component": "Cooling Fan",
        "severity": "Sev4",
        "service": "Cooling remains active, but the fan cycle trend indicates a need for maintenance review.",
        "technician": "Check for airflow restrictions and lubrication issues before the next peak demand period.",
    },
]

critical_templates = [
    {
        "title": "Payment Blocked",
        "description": "The cashless payment system is unable to authorize any transactions and the machine is blocked.",
        "component": "Payment Terminal",
        "severity": "Sev1",
        "service": "The machine cannot accept cashless payments and must be repaired before service resumes.",
        "technician": "Inspect the card reader hardware, verify all wiring, and replace the terminal if the self-test fails.",
    },
    {
        "title": "Boiler Overheat",
        "description": "The boiler has exceeded the safe operating temperature and entered a protective shutdown state.",
        "component": "Boiler and Heating Circuit",
        "severity": "Sev1",
        "service": "Hot beverage service is suspended until the thermal fault is cleared and the heater control is validated.",
        "technician": "Power down the unit, verify thermostat and thermal fuse operation, and replace failed heating components.",
    },
    {
        "title": "Water Supply Lost",
        "description": "The machine cannot complete beverage preparation because the water supply is unavailable.",
        "component": "Water Reservoir",
        "severity": "Sev2",
        "service": "Beverage service is unavailable until the system is refilled and the inlet path is verified.",
        "technician": "Refill the reservoir, inspect the inlet valve, and confirm the pump starts normally under load.",
    },
    {
        "title": "Controller Fault",
        "description": "The main controller failed to complete startup initialization and entered a hard fault state.",
        "component": "Main Controller",
        "severity": "Sev1",
        "service": "The unit cannot serve customers until the controller has been restarted and verified.",
        "technician": "Reboot the controller, reflash firmware if needed, and validate all sensor states before returning service.",
    },
    {
        "title": "Grinder Jammed",
        "description": "The grinder is blocked and all beverage dispensing operations are stopped.",
        "component": "Coffee Grinder",
        "severity": "Sev2",
        "service": "No beverage preparation can continue until the grinder is unclogged and tested.",
        "technician": "Clear the jam, inspect the grind chamber, and verify the assembly runs normally through a calibration cycle.",
    },
    {
        "title": "Boiler Leak Detected",
        "description": "A critical leak was detected in the boiler chamber and the machine was safely shut down.",
        "component": "Boiler Chamber",
        "severity": "Sev1",
        "service": "The machine must remain offline until the chamber and seals have been inspected and repaired.",
        "technician": "Isolate the water supply, inspect the chamber seal and body, and replace any failed components before restart.",
    },
    {
        "title": "Milk Pump Failed",
        "description": "The milk pumping circuit has failed and dairy-based beverages are unavailable.",
        "component": "Milk Pump",
        "severity": "Sev2",
        "service": "Milk beverages cannot be served until the pump is replaced and the circuit is primed.",
        "technician": "Replace the pump assembly, prime the line, and validate flow across three test cycles.",
    },
    {
        "title": "Power Supply Failure",
        "description": "The primary power supply is unstable and is interrupting machine operation.",
        "component": "Power Supply Unit",
        "severity": "Sev1",
        "service": "The machine has dropped functions and must remain offline until the PSU is replaced.",
        "technician": "Measure output voltage under load, replace the failing supply, and verify no downstream board is experiencing brown-out conditions.",
    },
    {
        "title": "Door Lockout Engaged",
        "description": "A door safety lockout has engaged, preventing all service operations from resuming.",
        "component": "Door Interlock",
        "severity": "Sev1",
        "service": "The machine is locked and cannot operate until the latch state is corrected and reset.",
        "technician": "Inspect the actuator, verify the sensor wiring, and clear the lockout before restoring service.",
    },
    {
        "title": "Dispense Pump Cavitation",
        "description": "The dispenser pump is cavitating and no longer delivers a reliable beverage flow.",
        "component": "Dispense Pump",
        "severity": "Sev1",
        "service": "Beverage service is stopped until the pump is primed or replaced and tested.",
        "technician": "Prime the pump, inspect for air ingress or blockage, and replace the pump if pressure recovery is not achieved.",
    },
]

alerts = []
for idx in range(19):
    template = info_templates[idx % len(info_templates)]
    alerts.append({
        "type": "Informational",
        "severity": "Sev5",
        "title": template["title"],
        "description": template["description"],
        "component": template["component"],
        "collection": taxonomy_map.get(template["component"], "Control"),
        "service": template["service"],
        "technician": template["technician"],
    })
for idx in range(40):
    template = warning_templates[idx % len(warning_templates)]
    alerts.append({
        "type": "Warning",
        "severity": template["severity"],
        "title": template["title"],
        "description": template["description"],
        "component": template["component"],
        "collection": taxonomy_map.get(template["component"], "Control"),
        "service": template["service"],
        "technician": template["technician"],
    })
for idx in range(16):
    template = critical_templates[idx % len(critical_templates)]
    alerts.append({
        "type": "Critical",
        "severity": template["severity"],
        "title": template["title"],
        "description": template["description"],
        "component": template["component"],
        "collection": taxonomy_map.get(template["component"], "Control"),
        "service": template["service"],
        "technician": template["technician"],
    })

headers = [
    "ID",
    "Alert Title",
    "Type",
    "Severity",
    "Alert Description",
    "Component",
    "Operator Response",
    "Model",
    "Last Update",
    "Version",
    "Notes",
    "Critical Stop Response",
    "Service Response",
    "Technician Response",
]

operator_responses = {
    "Maintenance Reminder Pending": "Run service restart.\nRun maintenance cycle.",
    "Low Bean Hopper": "Refill bean hopper.",
    "Milk Tank Below Target": "Refill milk tank.",
    "Water Level Warning": "Refill water tank.",
    "Payment Retry Rate Elevated": "Disconnect and reconnect payment machine.\nTry with secondary payment method.\nReconnect primary machine.",
    "Drip Tray Almost Full": "Empty drip tray.",
    "Payment Blocked": "Check network is up.\nFollow payment machine restart procedure.\nRetry again.\nIf problem persists, call service team.",
    "Boiler Overheat": "Wait for 15 minutes.\nTry again.",
    "Controller Fault": "Follow machine restart procedure.\nIf problem persists, call service team.",
    "Grinder Jammed": "Empty coffee grinder.\nClean hopper.\nRestart machine.\nTry again.\nIf issue persists, call service team.",
    "Boiler Leak Detected": "Turn off machine.\nEmpty tank.\nCall service team.",
    "Power Supply Failure": "Do not use machine.\nTurn off mains power.\nCall service team.",
    "Door Lockout Engaged": "Check door latch.\nCheck for occlusions.\nRestart machine.",
    "Dispense Pump Cavitation": "Turn off machine.\nRun coffee maintenance cycle twice.\nRetry.\nIf problem persists, call service team.",
}

models = ("All", "SC1x, SC3x", "SC2x", "SC2x, SC3x")
column_widths = {
    "A": 8.16, "B": 31, "C": 18, "E": 41.66, "F": 13.66,
    "G": 27.66, "H": 13, "I": 14.16, "K": 17.33, "L": 18,
    "M": 32.66, "N": 29.66,
}


def _row_height(values):
    """Estimate Excel's wrapped row height, capped like the reference workbook."""
    widths = [column_widths.get(get_column_letter(i), 8.83) for i in range(1, 15)]
    line_count = 1
    for value, width in zip(values, widths):
        text = "" if value is None else str(value)
        wrapped = sum(max(1, (len(line) + max(1, int(width)) - 1) // max(1, int(width))) for line in text.split("\n"))
        line_count = max(line_count, wrapped)
    return min(96, max(48, line_count * 16))


def build_workbook():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Alerts"
    worksheet.append(headers)

    first_update = date(2023, 12, 15)
    for index, alert in enumerate(alerts, start=1):
        last_update = first_update + timedelta(days=index - 1)
        version_year = 2025 if alert["type"] == "Critical" else 2023
        row = [
            str(index),
            alert["title"],
            alert["type"],
            alert["severity"],
            alert["description"],
            alert["component"],
            operator_responses.get(alert["title"], ""),
            models[(index - 1) // 19 % len(models)],
            last_update,
            f"DT 24.{version_year}.{index:02d}",
            "",
            "Yes" if alert["type"] == "Critical" else "",
            alert["service"],
            alert["technician"],
        ]
        worksheet.append(row)
        worksheet.row_dimensions[worksheet.max_row].height = _row_height(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=14):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet["A1"].number_format = "@"
    for cell in worksheet["A"][1:]:
        cell.number_format = "@"
    for cell in worksheet["I"][1:]:
        cell.number_format = "d-mmm-yy"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:N{worksheet.max_row}"
    worksheet.sheet_view.zoomScale = 150
    return workbook


def main():
    parser = argparse.ArgumentParser(description="Generate the cashless machine alert catalogue.")
    parser.add_argument("-o", "--output", default="cashless_machine_alerts.xlsx", help="Output .xlsx path")
    args = parser.parse_args()
    output = Path(args.output)
    if "donotdelete" in output.name.lower():
        parser.error("refusing to overwrite a DONOTDELETE reference workbook")
    if output.suffix.lower() != ".xlsx":
        parser.error("output must use the .xlsx extension")
    output.parent.mkdir(parents=True, exist_ok=True)
    build_workbook().save(output)
    print(f"Created {output} with {len(alerts)} alerts.")


if __name__ == "__main__":
    main()
